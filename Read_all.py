import os
from os.path import isdir

from matplotlib import rcParams
from openpyxl import load_workbook

import pandas as pd
import matplotlib.pyplot as plt
import win32com.client
import warnings

from datetime import datetime
from colorama import init, Fore, Style


def make(file_path):
    wb = load_workbook(file_path, data_only=True)

    sheet_names = wb.sheetnames

    if "prosinec" in sheet_names:
        last_sheet = wb["prosinec"]
        value = last_sheet["A11"].value

    elif "december" in sheet_names:
        last_sheet = wb["december"]
        value = last_sheet["A11"].value

    elif "4.q" in  sheet_names:
        last_sheet = wb["4.q"]
        value = last_sheet["A11"].value

    elif "4.Q" in sheet_names:
        last_sheet = wb["4.Q"]
        value = last_sheet["A11"].value

    elif "4.q 2025" in  sheet_names:
        last_sheet = wb["4.q 2025"]
        value = last_sheet["A11"].value

    elif "4.Q 2025" in sheet_names:
        last_sheet = wb["4.Q 2025"]
        value = last_sheet["A11"].value

    elif "zav" in sheet_names or "záv" in sheet_names:
        #TODO
        print("SOM TU")
        return 0
    else:
        print(Fore.RED + "Nenajdene v" + file_path)
        return 0

    print(Fore.GREEN + value)
    return  value


def main():
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")
    init(autoreset=True)
    folder = r"C:\Users\dominik.pozgay\OneDrive - ADMIO s.r.o\FVL"
    #folder = r"C:\Users\dominik.pozgay\OneDrive - ADMIO s.r.o\ADMIO - FV_DL"

    names = []

    for sub_folder in os.listdir(folder):
        if not isdir(folder + '\\' + sub_folder):
            continue

        for file in os.listdir(folder + '\\' + sub_folder):
            if "Dodací list 2025" in file:
                val = make(folder + '\\' + sub_folder + '\\' + file)
                if val != 0:
                    names.append(val)

    print(names)

    new_touples = {}

    with open("companies.txt", "r") as text:
        for line in text:
            split_line = line.split("\t")
            print(split_line)
            id = split_line[0]
            actual_name = split_line[2]
            for new in names:
                new_clean = new
                for trash in [".", " ", "-", "_"]:
                    new_clean = new_clean.replace(trash, "")

                old_clean = actual_name
                for trash in [".", " ", "-", "_"]:
                    old_clean = old_clean.replace(trash, "")

                print(old_clean.lower(), new_clean.lower(), old_clean.lower() in new_clean.lower())
                if old_clean.lower() in new_clean.lower():
                    print(Fore.GREEN + "OK")
                    # print(old_clean.lower(), new_clean)
                    new_touples[id] = new

    for i in range(1, 200):
        #print(str(i), new_touples.get(str(i)))
        name = new_touples.get(str(i))
        if name is not None:
            print(f"SET full_name = '{new_touples.get(str(i))}' WHERE company_id = {i}")
        else:
            print(f"SET isActive = 0 WHERE company_id = {i}")



    return 0


if __name__ == '__main__':
    # with open('sql.txt', 'r') as text:
    #     for line in text:
    #         print('UPDATE companies_invoicing')
    #         print(line.strip(), ";")
    main()